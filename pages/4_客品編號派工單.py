import streamlit as st
import pandas as pd
import datetime
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import io

st.header(" 客品編號派工單填寫系統（含下載）")

# ========= 檢查 NAS 連線 =========
def check_nas_connection():
    """檢查 NAS 連線狀態"""
    try:
        # 取得今天的日期
        today = datetime.date.today()
        roc_year = today.year - 1911
        filename = f"客品管理總表{roc_year}.{today.month:02d}.{today.day:02d}.xlsx"
        
        # NAS 路徑
        nas_path = fr"\\192.168.1.130\nas\倉管\{filename}"
        
        # 檢查檔案是否存在
        if os.path.exists(nas_path):
            return True, nas_path, filename
        else:
            return False, nas_path, filename
    except Exception as e:
        return False, None, None

# 檢查連線狀態
nas_connected, nas_path, filename = check_nas_connection()

if not nas_connected:
    st.error("⚠️ **無法連接到 NAS 網路磁碟機**")
    st.warning("""
    **可能的原因：**
    - 網路連線中斷
    - NAS 伺服器離線
    - 權限不足
    - 檔案路徑錯誤
    
    **請檢查：**
    1. 網路連線是否正常
    2. NAS 伺服器是否運作中
    3. 檔案路徑：`\\\\192.168.1.130\\nas\\倉管\\{filename}`
    """)
    st.stop()

# 顯示連線成功訊息
st.success(f"✅ 已成功連接到 NAS，使用檔案：`{filename}`")

# ========= 手動輸入 =========
target_code = st.text_input("請輸入客品編號（從總表中查詢）")

if target_code:
    try:
        # === 讀取客品管理總表 ===
        df_main = pd.read_excel(nas_path)

        # === 查詢客品資料 ===
        match_row = df_main[df_main['Unnamed: 1'] == target_code]
        if match_row.empty:
            st.error(" 找不到對應的客品編號")
        else:
            matched = match_row.iloc[0]
            品號 = str(matched['Unnamed: 1'])
            型號 = str(matched['Unnamed: 5'])
            salesperson = matched['Unnamed: 3']
            customer = matched['Unnamed: 4']
            item_code = 品號 + 型號
            today_str = datetime.datetime.today().strftime("%Y-%m-%d")

            # === 讀取範本檔案 ===
            dispatch_path = r"C:\Users\ASUS-PC47\Desktop\測試派工單.xlsx"
            wb = load_workbook(dispatch_path)
            ws = wb.active

            # === 公司別處理 ===
            company_raw = str(ws["A1"].value)
            if "德" in company_raw:
                company_label = "█德烜 □西得"
            elif "西" in company_raw:
                company_label = "□德烜 █西得"
            else:
                company_label = "□德烜 □西得"

            def safe_write(cell, value):
                if not isinstance(ws[cell], MergedCell):
                    ws[cell] = value

            # === 寫入內容 ===
            safe_write("C3", today_str)
            safe_write("A7", item_code)
            safe_write("A1", company_label)
            safe_write("G3", salesperson)
            safe_write("K3", customer)
            safe_write("F7", 1)
            safe_write("L7", 1)

            # === 將 Excel 存入記憶體 BytesIO ===
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.success(" 已成功填入資料！請點下方下載 Excel：")
            st.download_button(
                label=" 下載派工單 Excel",
                data=output,
                file_name="派工單_輸出.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"⚠️ 發生錯誤：{e}")
        st.error("請確認 NAS 連線狀態和檔案路徑")
