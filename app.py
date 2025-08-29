import streamlit as st
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import io

st.title(" 客品編號派工單填寫系統（含下載）")

# ========= 手動輸入 =========
target_code = st.text_input("請輸入客品編號（從總表中查詢）")

if target_code:
    try:
        # === 讀取客品管理總表 ===
        main_path = r"C:\Users\ASUS-PC47\Desktop\客品管理總表114.07.28.xlsx"
        df_main = pd.read_excel(main_path)

        # === 查詢客品資料 ===
        match_row = df_main[df_main['Unnamed: 1'] == target_code]
        if match_row.empty:
            st.error(" 找不到對應的客品編號")
        else:
            matched = match_row.iloc[0]
            品號 = str(matched['Unnamed: 1'])
            型號 = str(matched['Unnamed: 5'])
            salesperson = matched['Unnamed: 3']
            customer = matched['Unnamed: 4']
            item_code = 品號 + 型號
            today_str = datetime.datetime.today().strftime("%Y-%m-%d")

            # === 讀取範本檔案 ===
            dispatch_path = r"C:\Users\ASUS-PC47\Desktop\測試派工單.xlsx"
            wb = load_workbook(dispatch_path)
            ws = wb.active

            # === 公司別處理 ===
            company_raw = str(ws["A1"].value)
            if "德" in company_raw:
                company_label = "█德烜 □西得"
            elif "西" in company_raw:
                company_label = "□德烜 █西得"
            else:
                company_label = "□德烜 □西得"

            def safe_write(cell, value):
                if not isinstance(ws[cell], MergedCell):
                    ws[cell] = value

            # === 寫入內容 ===
            safe_write("C3", today_str)
            safe_write("A7", item_code)
            safe_write("A1", company_label)
            safe_write("G3", salesperson)
            safe_write("K3", customer)
            safe_write("F7", 1)
            safe_write("L7", 1)

            # === 將 Excel 存入記憶體 BytesIO ===
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.success(" 已成功填入資料！請點下方下載 Excel：")
            st.download_button(
                label=" 下載派工單 Excel",
                data=output,
                file_name="派工單_輸出.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"⚠️ 發生錯誤：{e}")
